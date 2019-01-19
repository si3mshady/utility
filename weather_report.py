import openpyxl
from weather_data import Brick_Tamland
from pprint import pprint   

class Spreadsheet:
    def __init__(self):
        self.excel_book = self.make_excel_wb()

    def make_excel_wb(self):
        return  openpyxl.Workbook()
    
    def get_sheetnames(self):       
       if len(self.excel_book.sheetnames) >= 2:       
            self.scribble_sheet()
       else:
            print('Please add sheets to your workbook. ')            
            self.create_sheet()
         
    def save_spreadsheet(self):
       self.save_sheet =  input("Save file as:> ")
       self.excel_book.save(self.save_sheet + '.xlsx')
       print('{} saved.'.format(self.save_sheet + '.xlsx'))
    
    def create_sheet(self):  
        self.name = input("Please provide a name for the worksheet:> ")      
        self.new_sheet = self.excel_book.create_sheet(index=len(self.excel_book.sheetnames) + 1 ,title=self.name)
        self.get_sheetnames()
    
    def scribble_sheet(self):       
        for sheetname in self.excel_book.sheetnames:
            print(">> " + sheetname + " <<")
        sheet  = input('Please enter a sheet to edit from the list. ')
        while sheet not in list(self.excel_book.sheetnames):
            sheet  = input('Please enter a sheet to edit from the list. ')
        self.scribble = self.excel_book[sheet]
        cuidad = input('Please enter a city name for a current weather report ')
        self.weather_man = Brick_Tamland(cuidad)
        currently = self.weather_man.weather_report                 
        row = self.scribble.max_row + 1
        self.scribble['A{}'.format(row)] = currently['location']['name']
        self.scribble['B{}'.format(row)] = currently['location']['localtime']
        self.scribble['C{}'.format(row)] = currently['current']['condition']['text']
        self.scribble['D{}'.format(row)] = currently['current']['feelslike_c']
        self.scribble['E{}'.format(row)] = currently['current']['feelslike_f']                       

def weather_report():
    print('This program will retrieve and insert weather data into spreadsheet workbooks.')
    go = True
    weather_man = Spreadsheet()
    while go:        
        weather_man.get_sheetnames()
        answer = input('Would you like to continue? (\'y\'\\\'n\')')
        if answer.lower() == 'n':
            print('Goodbye')
            weather_man.save_spreadsheet()
            go = False 
    
if __name__ == "__main__":
    weather_report()
 
 #Python 3 diy exercies; Inserting data into spreadsheets retrieved from API sources late-night toil, burning the midnight oil  Elliott Arnold 1-19-19

               
            


      
        