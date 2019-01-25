import openpyxl,re
from urllib.request import urlopen as URL
from bs4 import BeautifulSoup as soup

class Spreadsheet:
    def __init__(self):
        self.excel_book = self.make_excel_wb() 
        self.active = None      
        
    def make_excel_wb(self):
        return  openpyxl.Workbook()    

    def save_spreadsheet(self):
       self.save_sheet =  input("Save file as:> ")
       self.excel_book.save(self.save_sheet + '.xlsx')
       print('{} saved.'.format(self.save_sheet + '.xlsx'))
    
    def create_sheet(self):  
        self.name = input("Please provide a name for the worksheet:> ")      
        self.active = self.excel_book.create_sheet(index=len(self.excel_book.sheetnames) + 1 ,title=self.name)
               
    def scribble_sheet(self,vuln,description,index):                       
        self.row = index 
        self.active['A{}'.format(self.row)] = vuln 
        self.active['B{}'.format(self.row)] = description[:40]               
                              
def fetch_vulns():
    vSheet = Spreadsheet()
    vSheet.create_sheet()
    urlbase = "https://cve.mitre.org/cgi-bin/cvekey.cgi?keyword="
    masterP = '''[^C][^V][^E]p">([A-Z]*[a-z\s\w\.,\(\)\-:*"\$&;'#\/]+)''' #masterParser
    cP = '''>(CVE[-\d{4,10}]+)<'''	#cveParser
    searchString = input("Please enter keyword to search Common Vulnerabilities and Exposures\n")
    mainURL = urlbase + searchString
    raw_html = URL(mainURL)
    htmlSoup = soup(raw_html,"html.parser")
    processedSoup = htmlSoup.findAll('td', valign='top')
    cve = re.findall(r'>(CVE[-\d{4,10}]+)<',str(processedSoup))
    descriptions = re.findall(r'[^C][^V][^E]p">([A-Z]*[a-z\s\w\.,\(\)\-:*"\$&;\'#\/]+)',str(processedSoup))
    result = dict(zip(cve,descriptions))
    index = 1
    for key,value in result.items():
        vSheet.scribble_sheet(key,value,index)
        index += 1 
    vSheet.save_spreadsheet()               

if __name__ == "__main__":
    fetch_vulns()
#Enhanced GotVulns script that now parses cve data into spreadsheets = diy practice  1-24-19    Elliott Arnold = Si3mshady        
            

